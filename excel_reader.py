import openpyxl
import json
from entity import *
from myencoder import MyEncoder


class ExcelReader:

    wb = None
    sheets = None
    list = []
    dict = {}
    options = None

    def __init__(self, filename, options):
        self.wb = openpyxl.load_workbook(filename)
        self.options = options
        sheets = self.wb.sheetnames
        print(sheets, type(sheets))

    def read_sheet(self, sheet_name):
        if self.wb is None:
            return

        option = self.options[sheet_name]
        ws = self.wb.get_sheet_by_name(sheet_name)
        print(ws.max_column, ws.max_row)
        entity_list = []
        icol = 1
        department = ""
        last_position = ""
        while icol < ws.max_column:
            for irow in range(ws.max_row):
                if self.is_department(irow+1, icol, sheet_name):
                    department = self.get_range_value(irow+1, icol, sheet_name)
                    continue
                position = self.get_value(irow+1, icol+option["position"], sheet_name)
                name = ws.cell(irow+1, icol+option["name"]).value
                tel = ws.cell(irow+1, icol+option["tel"]).value
                mobile = ws.cell(irow+1, icol+option["mobile"]).value
                office = ws.cell(irow+1, icol+option["office"]).value
                company = sheet_name
                if position != last_position:
                    leader = True
                    last_position = position
                else:
                    leader = False
                # print(department, position, name, tel, mobile, office)

                if self.is_valid_name(name) is False:
                    continue
                elif self.is_valid(department) == self.is_valid(position) == self.is_valid(tel) == self.is_valid(mobile) == self.is_valid(office) is False:
                    continue
                else:
                    e = Entity(department, position, name, tel, mobile, office, leader, company)
                    #e.clear()
                    entity_list.append(e)
            icol += 5
        self.dict[sheet_name] = entity_list

    def get_range_value(self, row, col, sheet_name):
        # 获取指定的表单
        ws = self.wb.get_sheet_by_name(sheet_name)  # '研发中心'
        ranges = ws.merged_cells
        for rg in ranges:
            if rg.min_row <= row <= rg.max_row and rg.min_col <= col <= rg.max_col:
                return ws.cell(rg.min_row, rg.min_col).value
        return None

    def get_value(self, row, col, sheet_name):
        ws = self.wb.get_sheet_by_name(sheet_name)
        if ws.cell(row, col).value is not None:
            return ws.cell(row, col).value
        else:
            ranges = ws.merged_cells
            for rg in ranges:
                if rg.min_row <= row <= rg.max_row and rg.min_col <= col <= rg.max_col:
                    return ws.cell(rg.min_row, rg.min_col).value
        return None

    def is_department(self, row, col, sheet_name):
        ws = self.wb.get_sheet_by_name(sheet_name)
        ranges = ws.merged_cells
        for rg in ranges:
            if rg.max_col - rg.min_col == self.options[sheet_name]["col_count"]-1 and rg.min_col <= col <= rg.max_col and rg.max_row == rg.min_row == row:
                return True
        return False

    def is_valid(self, value):
        if value is None or value == "":
            return False
        else:
            return True

    def is_valid_name(self, name):
        print(isinstance(name,str))
        print(isinstance(name,int))
        if name is None:
            return  False
        if name == "":
            return False
        if isinstance(name,str) == False:
            return False
        return True

    def write_file(self, filename):
        sheets = self.wb.sheetnames

        with open(filename,'w', encoding='utf-8') as file_obj:
            for sheet_name in sheets:
                if self.dict[sheet_name] is not None and len(self.dict[sheet_name]) > 0:
                #if sheet_name == "北京研发中心" or sheet_name == "上海数据中心" or sheet_name == "股份公司":
                    json.dump(self.dict[sheet_name], file_obj, cls=MyEncoder,ensure_ascii=False,indent=4)
